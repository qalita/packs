"""
# QALITA (c) COPYRIGHT 2025 - ALL RIGHTS RESERVED -

Data compare — streaming implementation.

The comparison is a single lazy full join between the two sources. Row counts
and per-column mismatch counts come out of one streamed aggregation over that
join, so the only rows that ever reach memory are the bounded mismatch
examples.

Two properties of the previous version are gone on purpose:

* it loaded both sides fully into pandas and handed them to
  ``datacompy.Compare``, whose outer merge materializes a frame holding every
  column of both sides;
* above a row threshold it ``head()``-sampled each side INDEPENDENTLY. Head
  sampling two sides of a join yields near-disjoint key sets, so the reported
  precision / recall / f1 were wrong rather than approximate.

``datacompy`` is no longer a dependency: ``datacompy.PolarsCompare`` (0.18.1)
rejects LazyFrames outright — ``_validate_dataframe`` raises
``TypeError(f"{index} must be a Polars DataFrame")`` — and its join is an eager
``df1.join(df2, how="full", ...)``, so it cannot be put on the big path.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Mapping, Sequence, Tuple

import polars as pl

from qalita_core import analytics
from qalita_core.pack import Pack

logger = logging.getLogger(__name__)

# Side markers. A full join with coalesce=True merges the key columns, so the
# only way left to tell which side a row came from is an explicit marker.
_IN_SOURCE = "__qalita_in_source"
_IN_TARGET = "__qalita_in_target"

# Suffix given to the target copy of a compared column.
DEFAULT_TARGET_SUFFIX = "_target"
FALLBACK_TARGET_SUFFIX = "__qalita_target"

# Bounded mismatch examples. The metric carries evidence, not a data extract.
DEFAULT_MISMATCH_EXAMPLES = 10
MAX_MISMATCH_EXAMPLES = 1_000


def resolve_compare_columns(
    source_schema: Mapping[str, Any],
    target_schema: Mapping[str, Any],
    compare_col_list: Sequence[str],
    id_columns: Sequence[str],
    *,
    source_label: str = "source",
    target_label: str = "target",
) -> Tuple[List[str], List[str]]:
    """Columns to compare and join keys.

    Both come from the parquet footers, so choosing them costs no data read.
    """
    if compare_col_list:
        use_cols = list(dict.fromkeys(compare_col_list))
    else:
        # Keep the source order: a set intersection would make the column list,
        # and therefore every derived metric key, non-deterministic.
        use_cols = [c for c in source_schema if c in target_schema]

    missing_in_source = [c for c in use_cols if c not in source_schema]
    missing_in_target = [c for c in use_cols if c not in target_schema]
    if missing_in_source:
        raise ValueError(
            f"Columns missing in source {source_label}: {missing_in_source}"
        )
    if missing_in_target:
        raise ValueError(
            f"Columns missing in target {target_label}: {missing_in_target}"
        )

    keys = list(dict.fromkeys(id_columns)) or list(use_cols)
    missing_keys = [
        c for c in keys if c not in source_schema or c not in target_schema
    ]
    if missing_keys:
        raise ValueError(f"id_columns missing on one side: {missing_keys}")

    columns = list(dict.fromkeys(list(use_cols) + keys))
    return columns, keys


def target_suffix(columns: Sequence[str]) -> str:
    """A suffix that cannot collide with an existing column name.

    Any column already ending in the default suffix is enough to make the
    renaming ambiguous, so the check is on the ending rather than on an exact
    collision.
    """
    if any(str(c).endswith(DEFAULT_TARGET_SUFFIX) for c in columns):
        return FALLBACK_TARGET_SUFFIX
    return DEFAULT_TARGET_SUFFIX


def build_join(
    source_lf: "pl.LazyFrame",
    target_lf: "pl.LazyFrame",
    columns: Sequence[str],
    id_columns: Sequence[str],
    suffix: str,
) -> "pl.LazyFrame":
    """Full outer join of both sides on the join keys, as a lazy plan.

    ``nulls_equal`` matches datacompy's ``join_nulls=True``: a null key is a
    value like any other here, otherwise every null-keyed row would be reported
    as present on one side only.
    """
    source = source_lf.select(columns).with_columns(
        pl.lit(True).alias(_IN_SOURCE)
    )
    target = target_lf.select(columns).with_columns(
        pl.lit(True).alias(_IN_TARGET)
    )
    return source.join(
        target,
        on=list(id_columns),
        how="full",
        coalesce=True,
        suffix=suffix,
        nulls_equal=True,
    )


def unequal_expr(
    column: str,
    source_dtype: Any,
    target_dtype: Any,
    abs_tol: float,
    rel_tol: float,
    suffix: str,
) -> "pl.Expr":
    """True when the two sides disagree on ``column``.

    Two nulls agree, one null disagrees. Numeric columns use datacompy's
    tolerance rule, ``|a - b| <= abs_tol + rel_tol * |b|``. Columns whose dtypes
    differ are compared as text rather than raising, which is what makes a
    schema drift show up as mismatching values instead of a crash.
    """
    left = pl.col(column)
    right = pl.col(f"{column}{suffix}")

    both_numeric = bool(
        getattr(source_dtype, "is_numeric", lambda: False)()
        and getattr(target_dtype, "is_numeric", lambda: False)()
    )
    if both_numeric:
        left_f = left.cast(pl.Float64)
        right_f = right.cast(pl.Float64)
        close = (left_f - right_f).abs() <= (abs_tol + rel_tol * right_f.abs())
        return (
            pl.when(left.is_null() & right.is_null())
            .then(pl.lit(False))
            .when(left.is_null() | right.is_null())
            .then(pl.lit(True))
            .otherwise(~close.fill_null(False))
        )

    if source_dtype != target_dtype:
        return left.cast(pl.String).ne_missing(right.cast(pl.String))
    return left.ne_missing(right)


def _in_common() -> "pl.Expr":
    return pl.col(_IN_SOURCE).fill_null(False) & pl.col(_IN_TARGET).fill_null(
        False
    )


def compare_stats(
    joined: "pl.LazyFrame",
    value_columns: Sequence[str],
    source_schema: Mapping[str, Any],
    target_schema: Mapping[str, Any],
    abs_tol: float,
    rel_tol: float,
    suffix: str,
) -> Dict[str, Any]:
    """Every count the pack reports, from ONE streaming pass over the join.

    Batching the per-column mismatch expressions here is the difference between
    one scan and one scan per compared column.
    """
    in_common = _in_common()
    in_source = pl.col(_IN_SOURCE).fill_null(False)
    in_target = pl.col(_IN_TARGET).fill_null(False)

    unequal = {
        column: unequal_expr(
            column,
            source_schema[column],
            target_schema[column],
            abs_tol,
            rel_tol,
            suffix,
        )
        for column in value_columns
    }

    exprs: Dict[str, "pl.Expr"] = {
        "rows_in_common": in_common.sum(),
        "source_only": (in_source & ~in_target).sum(),
        "target_only": (~in_source & in_target).sum(),
    }
    for index, column in enumerate(value_columns):
        exprs[f"unequal|{index}"] = (unequal[column] & in_common).sum()
    if value_columns:
        exprs["unequal_rows"] = (
            pl.any_horizontal(*[unequal[c] for c in value_columns]) & in_common
        ).sum()

    result = analytics.agg(joined, exprs)

    by_column = {
        column: int(result.get(f"unequal|{index}") or 0)
        for index, column in enumerate(value_columns)
    }
    return {
        "rows_in_common": int(result.get("rows_in_common") or 0),
        "source_only": int(result.get("source_only") or 0),
        "target_only": int(result.get("target_only") or 0),
        "unequal_rows": int(result.get("unequal_rows") or 0),
        "unequal_by_column": by_column,
        "unequal_values": int(sum(by_column.values())),
    }


def mismatch_examples(
    joined: "pl.LazyFrame",
    mismatching_columns: Sequence[str],
    id_columns: Sequence[str],
    source_schema: Mapping[str, Any],
    target_schema: Mapping[str, Any],
    abs_tol: float,
    rel_tol: float,
    suffix: str,
    limit: int,
) -> Tuple[int, "pl.DataFrame"]:
    """Bounded rows where at least one compared column disagrees.

    Only the columns that actually mismatch somewhere are kept, which is what
    ``ignore_matching_cols=True`` did before, minus the full frame it needed.
    """
    if not mismatching_columns or limit <= 0:
        return 0, pl.DataFrame()

    predicate = (
        pl.any_horizontal(
            *[
                unequal_expr(
                    column,
                    source_schema[column],
                    target_schema[column],
                    abs_tol,
                    rel_tol,
                    suffix,
                )
                for column in mismatching_columns
            ]
        )
        & _in_common()
    )
    columns = list(id_columns) + [
        name
        for column in mismatching_columns
        for name in (column, f"{column}{suffix}")
        if column not in id_columns
    ]
    count, rows = analytics.failures(
        joined, predicate, limit=limit, columns=columns
    )
    renamed = {
        name: (
            name
            if name in id_columns
            else (
                f"{name[: -len(suffix)]}_target"
                if name.endswith(suffix)
                else f"{name}_source"
            )
        )
        for name in rows.columns
    }
    return count, rows.rename(renamed)


def mismatches_table(
    rows: "pl.DataFrame", total_mismatches: int, limit: int
) -> Dict[str, Any]:
    """The table payload consumed by the UI chart."""
    labels = list(rows.columns)
    structure: Dict[str, Any] = {
        "columnLabels": labels,
        "data": [
            [{"value": value} for value in row] for row in rows.iter_rows()
        ],
    }
    if total_mismatches > limit:
        structure["truncated"] = True
        structure["total_mismatches"] = int(total_mismatches)
    return structure


def _normalize(key: str) -> str:
    return key.lower().replace(" ", "_")


def build_metrics(
    stats: Mapping[str, Any],
    source_label: str,
    target_label: str,
    columns: Sequence[str],
    source_schema: Mapping[str, Any],
    target_schema: Mapping[str, Any],
    source_rows: int,
    target_rows: int,
    abs_tol: float,
    rel_tol: float,
) -> List[Dict[str, Any]]:
    """The metric set, keeping the keys the previous report parser produced."""
    source_scope = {"perimeter": "dataset", "value": source_label}
    target_scope = {"perimeter": "dataset", "value": target_label}
    n_columns = len(columns)

    common_columns = [c for c in source_schema if c in target_schema]
    source_only_columns = [c for c in source_schema if c not in target_schema]
    target_only_columns = [c for c in target_schema if c not in source_schema]

    unequal_columns = sum(
        1 for count in stats["unequal_by_column"].values() if count > 0
    )
    compared_columns = len(stats["unequal_by_column"])

    metrics: List[Dict[str, Any]] = [
        {
            "key": f"dataframe_summary_number_columns_{source_label}",
            "value": n_columns,
            "scope": source_scope,
        },
        {
            "key": f"dataframe_summary_number_columns_{target_label}",
            "value": n_columns,
            "scope": target_scope,
        },
        {
            "key": f"dataframe_summary_number_rows_{source_label}",
            "value": int(source_rows),
            "scope": source_scope,
        },
        {
            "key": f"dataframe_summary_number_rows_{target_label}",
            "value": int(target_rows),
            "scope": target_scope,
        },
    ]

    parsed: Dict[str, Any] = {
        "column_summary_number_of_columns_in_common": len(common_columns),
        _normalize(
            f"column_summary_number_of_columns_in_{source_label}"
            f"_but_not_in_{target_label}"
        ): len(source_only_columns),
        _normalize(
            f"column_summary_number_of_columns_in_{target_label}"
            f"_but_not_in_{source_label}"
        ): len(target_only_columns),
        "row_summary_default_absolute_tolerance": abs_tol,
        "row_summary_default_relative_tolerance": rel_tol,
        "row_summary_number_of_rows_in_common": stats["rows_in_common"],
        _normalize(
            f"row_summary_number_of_rows_in_{source_label}"
            f"_but_not_in_{target_label}"
        ): stats["source_only"],
        _normalize(
            f"row_summary_number_of_rows_in_{target_label}"
            f"_but_not_in_{source_label}"
        ): stats["target_only"],
        "row_summary_number_of_rows_with_some_compared_columns_unequal": stats[
            "unequal_rows"
        ],
        "row_summary_number_of_rows_with_all_compared_columns_equal": max(
            stats["rows_in_common"] - stats["unequal_rows"], 0
        ),
        "column_comparison_number_of_columns_compared_with_some_values_unequal": (
            unequal_columns
        ),
        "column_comparison_number_of_columns_compared_with_all_values_equal": (
            compared_columns - unequal_columns
        ),
        "column_comparison_total_number_of_values_which_compare_unequal": (
            stats["unequal_values"]
        ),
    }
    metrics.extend(
        {"key": key, "value": str(value), "scope": source_scope}
        for key, value in parsed.items()
    )
    return metrics


def scores(
    stats: Mapping[str, Any], source_rows: int, target_rows: int
) -> Dict[str, float]:
    """Matching score, precision, recall and f1.

    These are only meaningful because both sides are now joined in full: the
    previous independent head() sampling gave near-disjoint key sets, so every
    number here was wrong rather than approximate.
    """
    common = stats["rows_in_common"]
    if target_rows > 0:
        score = max(0.0, 1.0 - (stats["unequal_rows"] / target_rows))
        precision = common / target_rows
    else:
        score = 0.0
        precision = 0.0
    recall = common / source_rows if source_rows > 0 else 0.0
    denominator = precision + recall
    f1 = (2 * precision * recall / denominator) if denominator else 0.0
    return {
        "score": score,
        "precision": precision,
        "recall": recall,
        "f1_score": f1,
    }


def build_report(
    source_label: str,
    target_label: str,
    stats: Mapping[str, Any],
    columns: Sequence[str],
    id_columns: Sequence[str],
    source_rows: int,
    target_rows: int,
    abs_tol: float,
    rel_tol: float,
) -> str:
    """Plain-text summary, in the shape the previous datacompy report had."""
    unequal = stats["unequal_by_column"]
    lines = [
        "QALITA Comparison",
        "-----------------",
        "",
        "DataFrame Summary",
        "-----------------",
        "",
        f"{source_label}: {len(columns)} columns, {source_rows} rows",
        f"{target_label}: {len(columns)} columns, {target_rows} rows",
        "",
        "Row Summary",
        "-----------",
        "",
        f"Matched on: {', '.join(id_columns)}",
        f"Default Absolute Tolerance: {abs_tol}",
        f"Default Relative Tolerance: {rel_tol}",
        f"Number of rows in common: {stats['rows_in_common']}",
        f"Number of rows in {source_label} but not in {target_label}: "
        f"{stats['source_only']}",
        f"Number of rows in {target_label} but not in {source_label}: "
        f"{stats['target_only']}",
        f"Number of rows with some compared columns unequal: "
        f"{stats['unequal_rows']}",
        f"Number of rows with all compared columns equal: "
        f"{max(stats['rows_in_common'] - stats['unequal_rows'], 0)}",
        "",
        "Column Comparison",
        "-----------------",
        "",
        f"Number of columns compared with some values unequal: "
        f"{sum(1 for n in unequal.values() if n > 0)}",
        f"Number of columns compared with all values equal: "
        f"{sum(1 for n in unequal.values() if n == 0)}",
        f"Total number of values which compare unequal: "
        f"{stats['unequal_values']}",
        "",
    ]
    mismatching = [c for c, n in unequal.items() if n > 0]
    if mismatching:
        lines.extend(["Columns with Unequal Values", "-" * 27, ""])
        lines.extend(f"{c}: {unequal[c]}" for c in mismatching)
        lines.append("")
    return "\n".join(lines)


def _examples_limit(pack_config: Mapping[str, Any]) -> int:
    job = pack_config.get("job") or {}
    raw = job.get("mismatch_examples", DEFAULT_MISMATCH_EXAMPLES)
    try:
        limit = int(raw)
    except (TypeError, ValueError):
        limit = DEFAULT_MISMATCH_EXAMPLES
    return max(0, min(limit, MAX_MISMATCH_EXAMPLES))


def write_excel(frame: "pl.DataFrame", path: str) -> str:
    """Write a bounded frame to xlsx, falling back to csv without openpyxl."""
    try:
        from openpyxl import Workbook
    except ImportError:  # pragma: no cover - depends on the install extras
        csv_path = os.path.splitext(path)[0] + ".csv"
        logger.warning(
            "openpyxl is not installed, writing %s instead of %s",
            csv_path,
            path,
        )
        frame.write_csv(csv_path)
        return csv_path

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("mismatches")
    sheet.append(list(frame.columns))
    for row in frame.iter_rows():
        sheet.append(
            [
                (
                    value
                    if isinstance(value, (int, float, str, type(None)))
                    else str(value)
                )
                for value in row
            ]
        )
    workbook.save(path)
    return path


def _pairings(
    source_tables: Sequence[str], target_tables: Sequence[str]
) -> List[Tuple[str, str]]:
    """Pair logical objects by position, on names rather than parquet parts.

    Pairing parquet paths is what used to drop chunks 2..N of a chunked table.
    """
    if len(source_tables) == len(target_tables):
        return list(zip(source_tables, target_tables))
    logger.warning(
        "Source/Target tables count mismatch (%s vs %s); comparing the first "
        "dataset of each.",
        len(source_tables),
        len(target_tables),
    )
    return [(source_tables[0], target_tables[0])]


def main() -> None:
    with Pack() as pack:
        if pack.source_config.get("type") == "database":
            table_or_query = pack.source_config.get("config", {}).get(
                "table_or_query"
            )
            if not table_or_query:
                raise ValueError(
                    "For a 'database' type source, you must specify "
                    "'table_or_query' in the config."
                )
            pack.load_data("source", table_or_query=table_or_query)
        else:
            pack.load_data("source")

        if pack.target_config.get("type") == "database":
            table_or_query = pack.target_config.get("config", {}).get(
                "table_or_query"
            )
            if not table_or_query:
                raise ValueError(
                    "For a 'database' type target, you must specify "
                    "'table_or_query' in the config."
                )
            pack.load_data("target", table_or_query=table_or_query)
        else:
            pack.load_data("target")

        job = pack.pack_config.get("job") or {}
        compare_col_list = job.get("compare_col_list", []) or []
        configured_ids = job.get("id_columns", []) or []
        abs_tol = float(job.get("abs_tol", 0.0001) or 0.0)
        rel_tol = float(job.get("rel_tol", 0) or 0.0)
        limit = _examples_limit(pack.pack_config)

        source_tables = pack.tables("source")
        target_tables = pack.tables("target")
        single = len(source_tables) == 1 and len(target_tables) == 1

        for s_table, t_table in _pairings(source_tables, target_tables):
            s_label = pack.source_config["name"] if single else s_table
            t_label = pack.target_config["name"] if single else t_table

            source_schema = pack.schema("source", s_table)
            target_schema = pack.schema("target", t_table)
            columns, id_columns = resolve_compare_columns(
                source_schema,
                target_schema,
                compare_col_list,
                configured_ids,
                source_label=s_label,
                target_label=t_label,
            )
            value_columns = [c for c in columns if c not in id_columns]
            suffix = target_suffix(columns)

            joined = build_join(
                pack.scan("source", s_table),
                pack.scan("target", t_table),
                columns,
                id_columns,
                suffix,
            )

            stats = compare_stats(
                joined,
                value_columns,
                source_schema,
                target_schema,
                abs_tol,
                rel_tol,
                suffix,
            )

            source_rows = pack.get_row_count("source", s_table)
            target_rows = pack.get_row_count("target", t_table)

            pack.metrics.data.extend(
                build_metrics(
                    stats,
                    s_label,
                    t_label,
                    columns,
                    source_schema,
                    target_schema,
                    source_rows,
                    target_rows,
                    abs_tol,
                    rel_tol,
                )
            )

            computed = scores(stats, source_rows, target_rows)
            logger.info(
                "[%s vs %s] score=%.4f precision=%.4f recall=%.4f f1=%.4f",
                s_label,
                t_label,
                computed["score"],
                computed["precision"],
                computed["recall"],
                computed["f1_score"],
            )
            pack.metrics.data.extend(
                {
                    "key": key,
                    "value": str(round(value, 2)),
                    "scope": {"perimeter": "dataset", "value": s_label},
                }
                for key, value in computed.items()
            )

            report = build_report(
                s_label,
                t_label,
                stats,
                columns,
                id_columns,
                source_rows,
                target_rows,
                abs_tol,
                rel_tol,
            )
            with open(
                f"comparison_report_{s_label}_vs_{t_label}.txt",
                "w",
                encoding="utf-8",
            ) as handle:
                handle.write(report)

            mismatching_columns = [
                column
                for column, count in stats["unequal_by_column"].items()
                if count > 0
            ]
            total_mismatches, rows = mismatch_examples(
                joined,
                mismatching_columns,
                id_columns,
                source_schema,
                target_schema,
                abs_tol,
                rel_tol,
                suffix,
                limit,
            )

            pack.metrics.data.extend(
                [
                    {
                        "key": "recommendation_levels_mismatches",
                        "value": {
                            "info": "0",
                            "warning": "0.5",
                            "high": "0.8",
                        },
                        "scope": {"perimeter": "dataset", "value": s_label},
                    },
                    {
                        "key": "mismatches_table",
                        "value": mismatches_table(
                            rows, total_mismatches, limit
                        ),
                        "scope": {"perimeter": "dataset", "value": s_label},
                    },
                ]
            )

            if rows.height and pack.source_config.get("type") == "file":
                source_file_dir = os.path.dirname(
                    pack.source_config["config"]["path"]
                )
                current_date = datetime.now().strftime("%Y%m%d")
                report_file_path = os.path.join(
                    source_file_dir,
                    f"{current_date}_data_compare_report_"
                    f"{s_label}_vs_{t_label}.xlsx",
                )
                written = write_excel(rows, report_file_path)
                logger.info(
                    "Mismatching rows have been exported to %s", written
                )

        pack.metrics.save()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()
