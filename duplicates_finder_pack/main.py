"""
# QALITA (c) COPYRIGHT 2025 - ALL RIGHTS RESERVED -

Duplicates finder — streaming implementation.

Everything this pack computes stays inside the Polars streaming engine: the
group-by that identifies duplicate key combinations never leaves the engine as
a Python object, and the only rows that reach memory are the bounded export
sample. The previous version read every parquet chunk into pandas before it
reached the (already streaming) counting helper, which made the counting helper
irrelevant to peak memory.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

import polars as pl

from qalita_core import analytics
from qalita_core.pack import Pack
from qalita_core.utils import determine_recommendation_level

logger = logging.getLogger(__name__)

# Name of the group size column. Prefixed so it cannot collide with a user
# column that happens to be called "count".
_COUNT = "__qalita_group_size"

# Bound on the rows written to the duplicates report. The report is a bounded
# artefact by construction: an unbounded export is what used to make a 100 GiB
# source unanalysable.
DEFAULT_DUPLICATE_ROWS_LIMIT = 10_000
MAX_DUPLICATE_ROWS_LIMIT = 100_000

# Score below which the pack raises a recommendation.
SCORE_RECOMMENDATION_THRESHOLD = 0.9


def resolve_uniqueness_columns(
    pack_config: Dict[str, Any], schema: Dict[str, Any]
) -> List[str]:
    """Columns whose combination defines a duplicate.

    Falls back to every column, which is the "identical rows" definition.
    Unknown configured columns are rejected rather than silently ignored: a
    typo there changes the meaning of the whole score.
    """
    configured = (pack_config.get("job") or {}).get(
        "compute_uniqueness_columns"
    ) or []
    if not configured:
        return list(schema)

    unknown = [col for col in configured if col not in schema]
    if unknown:
        raise ValueError(
            f"compute_uniqueness_columns refers to columns absent from the "
            f"dataset: {unknown}. Available: {sorted(schema)}"
        )
    return list(configured)


def duplicate_counts(
    lf: "pl.LazyFrame",
    uniqueness_columns: Sequence[str],
    *,
    exact: bool = True,
) -> Tuple[int, int, str]:
    """Total rows and duplicate rows, without materializing the group table.

    ``exact`` (the default here) groups by the uniqueness columns and sums
    ``size - 1`` inside the engine. The streaming engine spills group state to
    disk, so this holds on a source far larger than RAM.

    ``exact=False`` derives the duplicate count from a HyperLogLog distinct
    count. It is O(1) memory but it is a subtraction of two large numbers: at a
    low duplication rate the HLL error dwarfs the answer, so it is offered as an
    opt-in rather than the default that the rest of the suite uses.

    Returns:
        ``(total_rows, duplicates, method)``.
    """
    columns = list(uniqueness_columns)
    if not columns:
        raise ValueError("at least one uniqueness column is required")

    total_rows = analytics.row_count(lf)

    if exact:
        grouped = (
            lf.select(columns).group_by(columns).agg(pl.len().alias(_COUNT))
        )
        duplicates = analytics.agg(
            grouped,
            {"duplicates": (pl.col(_COUNT) - 1).clip(0).sum()},
        )["duplicates"]
        return total_rows, int(duplicates or 0), "exact"

    distinct = analytics.agg(
        lf,
        {"distinct": pl.struct(columns).hash().approx_n_unique()},
    )["distinct"]
    duplicates = max(total_rows - int(distinct or 0), 0)
    return total_rows, duplicates, "hyperloglog"


def duplicate_rows(
    lf: "pl.LazyFrame",
    uniqueness_columns: Sequence[str],
    *,
    limit: int,
) -> Tuple[int, "pl.DataFrame"]:
    """Rows that belong to a duplicated key group, exact count and bounded rows.

    The count comes from the engine; the rows go through
    :func:`analytics.failures`, which caps them inside the lazy plan so the cap
    holds however many rows are duplicated.
    """
    columns = list(lf.collect_schema().keys())
    keys = list(uniqueness_columns)

    counts = lf.group_by(keys).agg(pl.len().alias(_COUNT))
    # nulls_equal mirrors group_by, which puts null keys in a group of their
    # own; without it, duplicated all-null keys would be counted but never
    # exported.
    with_counts = lf.join(counts, on=keys, how="left", nulls_equal=True)

    return analytics.failures(
        with_counts,
        pl.col(_COUNT) > 1,
        limit=limit,
        columns=columns,
    )


def duplicate_metrics(
    dataset: str,
    total_rows: int,
    duplicates: int,
    method: str,
) -> List[Dict[str, Any]]:
    """Dataset-scoped metrics for one dataset.

    ``*_method`` siblings let the UI label a number as approximate instead of
    presenting a HyperLogLog estimate as a fact.
    """
    duplication_rate = duplicates / total_rows if total_rows > 0 else 0.0
    score = max(0.0, min(1.0, 1.0 - duplication_rate))
    distinct_count = max(total_rows - duplicates, 0)
    distinct_percent = distinct_count / total_rows if total_rows > 0 else 0.0
    scope = {"perimeter": "dataset", "value": dataset}

    return [
        {"key": "score", "value": str(round(score, 2)), "scope": scope},
        {"key": "duplicates", "value": int(duplicates), "scope": scope},
        {"key": "duplicates_method", "value": method, "scope": scope},
        {
            "key": "distinct_count",
            "value": int(distinct_count),
            "scope": scope,
        },
        {
            "key": "distinct_percent",
            "value": str(round(distinct_percent, 4)),
            "scope": scope,
        },
        {"key": "distinct_count_method", "value": method, "scope": scope},
        {"key": "rows", "value": int(total_rows), "scope": scope},
    ]


def _export_limit(pack_config: Dict[str, Any]) -> int:
    job = pack_config.get("job") or {}
    raw = job.get("duplicate_rows_limit", DEFAULT_DUPLICATE_ROWS_LIMIT)
    try:
        limit = int(raw)
    except (TypeError, ValueError):
        limit = DEFAULT_DUPLICATE_ROWS_LIMIT
    return max(0, min(limit, MAX_DUPLICATE_ROWS_LIMIT))


def write_excel(frame: "pl.DataFrame", path: str) -> str:
    """Write a bounded frame to xlsx, falling back to csv without openpyxl.

    Only ever called with a frame that is already capped, so the whole sheet is
    held in memory on purpose.
    """
    try:
        from openpyxl import Workbook
    except ImportError:  # pragma: no cover - depends on the install extras
        csv_path = os.path.splitext(path)[0] + ".csv"
        logger.warning(
            "openpyxl is not installed, writing %s instead of %s",
            csv_path,
            path,
        )
        frame.write_csv(csv_path)
        return csv_path

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("duplicates")
    sheet.append(list(frame.columns))
    for row in frame.iter_rows():
        sheet.append(
            [
                (
                    value
                    if isinstance(value, (int, float, str, type(None)))
                    else str(value)
                )
                for value in row
            ]
        )
    workbook.save(path)
    return path


def main() -> None:
    with Pack() as pack:
        if pack.source_config.get("type") == "database":
            table_or_query = pack.source_config.get("config", {}).get(
                "table_or_query"
            )
            if not table_or_query:
                raise ValueError(
                    "For a 'database' type source, you must specify "
                    "'table_or_query' in the config."
                )
            pack.load_data("source", table_or_query=table_or_query)
        else:
            pack.load_data("source")

        job = pack.pack_config.get("job") or {}
        # Exact by default: see duplicate_counts() for why the approximate
        # distinct count is a poor basis for a duplication rate.
        exact = bool(job.get("exact", True))
        limit = _export_limit(pack.pack_config)

        tables = pack.tables("source")
        # One logical object is one dataset. The parts of a chunked object are
        # already one LazyFrame, so there is nothing left to "treat as one".
        single = len(tables) == 1
        source_name = pack.source_config["name"]

        first_table: Optional[str] = None
        first_keys: List[str] = []

        for table in tables:
            dataset = source_name if single else table
            lf = pack.scan("source", table)
            schema = pack.schema("source", table)
            uniqueness_columns = resolve_uniqueness_columns(
                pack.pack_config, schema
            )
            logger.info(
                "[%s] checking duplicates on %s",
                dataset,
                uniqueness_columns,
            )

            total_rows, duplicates, method = duplicate_counts(
                lf, uniqueness_columns, exact=exact
            )
            logger.info(
                "[%s] %s duplicates out of %s rows (%s)",
                dataset,
                duplicates,
                total_rows,
                method,
            )

            pack.metrics.data.extend(
                duplicate_metrics(dataset, total_rows, duplicates, method)
            )

            if job.get("compute_uniqueness_columns"):
                # Legacy companion metric scoped to the uniqueness columns.
                pack.metrics.data.append(
                    {
                        "key": "duplicates",
                        "value": int(duplicates),
                        "scope": {
                            "perimeter": "dataset",
                            "value": ", ".join(uniqueness_columns),
                        },
                    }
                )

            duplication_rate = (
                duplicates / total_rows if total_rows > 0 else 0.0
            )
            score = max(0.0, min(1.0, 1.0 - duplication_rate))
            if score < SCORE_RECOMMENDATION_THRESHOLD:
                pack.recommendations.data.append(
                    {
                        "content": (
                            f"dataset '{dataset}' has a duplication rate of "
                            f"{duplication_rate * 100:.1f}% on the scope "
                            f"{list(uniqueness_columns)}."
                        ),
                        "type": "Duplicates",
                        "scope": {"perimeter": "dataset", "value": dataset},
                        "level": determine_recommendation_level(
                            duplication_rate
                        ),
                    }
                )

            if first_table is None:
                first_table = table
                first_keys = uniqueness_columns

        # Export the first dataset only, as before. The rows are bounded and
        # only materialized when a report will actually be written.
        if (
            first_table is not None
            and limit > 0
            and pack.source_config.get("type") == "file"
        ):
            dataset = source_name if single else first_table
            in_groups, rows = duplicate_rows(
                pack.scan("source", first_table), first_keys, limit=limit
            )
            pack.metrics.data.append(
                {
                    "key": "duplicated_rows",
                    "value": int(in_groups),
                    "scope": {"perimeter": "dataset", "value": dataset},
                }
            )
            if rows.height == 0:
                logger.info("No duplicates found. No report generated.")
            else:
                if in_groups > limit:
                    logger.info(
                        "Limiting duplicate export to %s of %s rows",
                        limit,
                        in_groups,
                    )
                source_file_dir = os.path.dirname(
                    pack.source_config["config"]["path"]
                )
                current_date = datetime.now().strftime("%Y%m%d")
                report_file_path = os.path.join(
                    source_file_dir,
                    f"{current_date}_duplicates_finder_report_"
                    f"{source_name}.xlsx",
                )
                written = write_excel(rows, report_file_path)
                logger.info(
                    "Duplicated rows have been exported to %s", written
                )

        pack.metrics.save()
        pack.recommendations.save()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()
